import matplotlib.pyplot as plt
import openpyxl
import numpy as np
import math

fig = plt.figure(figsize=(30, 10))
ax1 = fig.add_subplot(153)
ax2 = fig.add_subplot(155)
ax3 = fig.add_subplot(151)
'''fig = plt.figure(figsize=(16, 4.5))
ax1 = fig.add_subplot(132)
ax2 = fig.add_subplot(133)
ax3 = fig.add_subplot(131)'''

A = []
A1 = []
A3 = []#x для 3 графика
B = []
B3 = [] #y  для 3 графика
C = []
C1 = []
C3 = []
C4 = []
D = []
k = 0

book = openpyxl.load_workbook("Лаб_1_1_4.xlsx",data_only = True)#data_only считывает значение, а не формулу

sheet_1 = book.worksheets[0]
sheet_2 = book.worksheets[1]

for row in sheet_1.iter_rows(min_row=3, max_row=22, min_col=2, max_col=11):
    for cell in row:
        A.append(int(cell.value))

for i in range(min(A), max(A) + 1):
    B.append(i)

for y in range(min(A), max(A) + 1):
    for g in range(0, len(A)):
        if y == A[g]:
            k += 1
    C.append(k)
    k = 0

for i in range(0, len(A)):
    if i % 2 == 0:
        C1.append(A[i] + A[i + 1])

s = sum(C)
plt.ylim(0, 0.10)
for i in range(0, len(C)):
    C[i] = C[i] / s  # посчитана доля случаев

for row in sheet_1.iter_rows(min_row=2, max_row=2, min_col=1, max_col=21):
    for cell in row:
        A1.append(cell.value)

for col in range(1,21+1):
    A3.append(sheet_2[1][col].value)
    B3.append(sheet_2[3][col].value)


ax1.bar(B, C, width=1, color="skyblue", edgecolor="white", linewidth=0)

#plt.xlabel('n', fontsize=10, color='blue')
#plt.ylabel('w', fontsize=10, color='blue')

ax2.hist(C1, bins=max(C1) - min(C1), color='darkgreen', density=True)
# plt.hist(A, bins=28, density=True)
print(A3)
print(B3)
ax3.bar(A3, B3, width=1, color="purple", edgecolor="white", linewidth=0)

ax1.set_title('Гистограмма для 20 с', fontdict={'fontname': 'sans-serif', 'fontsize': 15})
ax2.set_title('Гистограмма для 40 с', fontdict={'fontname': 'sans-serif', 'fontsize': 15})
ax3.set_title('Гистограмма для 10 с', fontdict={'fontname': 'sans-serif', 'fontsize': 15})
# ax1.legend()
# ax2.legend()
ax1.set_xlabel('n, число импульсов', color='blue', fontdict=None, labelpad=None, loc=None)
ax1.set_ylabel('w, доля случаев', color='blue', fontdict=None, labelpad=None, loc=None)

ax2.set_xlabel('n, число импульсов', color='blue', fontdict=None, labelpad=None, loc=None)
ax2.set_ylabel('w, доля случаев', color='blue', fontdict=None, labelpad=None, loc=None)

ax3.set_xlabel('n, число импульсов', color='blue', fontdict=None, labelpad=None, loc=None)
ax3.set_ylabel('w, доля случаев', color='blue', fontdict=None, labelpad=None, loc=None)

ax1.set_ylim(0, 0.12)
ax2.set_ylim(0, 0.12)
ax3.set_ylim(0,0.12)

ax1.set_xlim(10, 42)
ax2.set_xlim(30,72)
ax3.set_xlim(0, 26)

###################__CALCULATIONS__########################################################
for i in range(len(A3)):
    D.append(A3[i]*B3[i]*400)
n_sr_1 = sum(D)/400

print(f'Среденее число срабатываний счетчика за 10с: n_sr_1 = {n_sr_1}')

for i in range (len(A3)):
    C3.append(B3[i]*400*(A3[i] - n_sr_1)**2)
s_1 = np.sqrt(1/400*sum(C3))

print(f'Среденеквадратичная ошибка s_1 = {s_1}')
print(f'Приблизительно s_1 = {np.sqrt(n_sr_1)}')

C3 = []

for i in range(len(A3)):
    if -s_1 + n_sr_1 <= A3[i] <= s_1 + n_sr_1:
        C3.append(B3[i]*400)
    if -2*s_1 + n_sr_1 <= A3[i] <= 2*s_1 + n_sr_1:
        C4.append(B3[i]*400)
print(f'Доля случаев+-s_1:{sum(C3)/400*100}%')
print(f'Доля случаев+-2*s_1:{sum(C4)/400*100}%')

C3 = []
C4 = []
D =[]
n_sr_2 = sum(A)/200
print(f'Среденее число срабатываний счетчика за 20с: n_sr_2 = {n_sr_2}')
for i in range (len(A)):
    C3.append((A[i] - n_sr_2)**2)
s_2 = np.sqrt(1/200*sum(C3))
print(f'Среденеквадратичная ошибка s_2 = {s_2}')
print(f'Приблизительно s_2 = {np.sqrt(n_sr_2)}')
for i in range(len(A)):
    if (-s_2 + n_sr_2 <= A[i] <= s_2+ n_sr_2):
        C4.append(1)
    if -2*s_2 + n_sr_2 <= A[i] <= 2*s_2 + n_sr_2:
        D.append(1)
print(f'Доля случаев+-s_2:{sum(C4)/200*100}%')
print(f'Доля случаев+-2*s_2:{sum(D)/200*100}%')
D = []
C3 = []
C4 = []

n_sr_3 = sum(C1)/100
print(f'Среденее число срабатываний счетчика за 40с: n_sr_3 = {n_sr_3}')
for i in range (len(C1)):
    C3.append((C1[i] - n_sr_3)**2)
s_3 = np.sqrt(1/100*sum(C3))
print(f'Среденеквадратичная ошибка s_3 = {s_3}')
print(f'Приблизительно s_3 = {np.sqrt(n_sr_3)}')

print(f's1/n_sr_1*100%={s_1/n_sr_1*100}%')
print(f's2/n_sr_2*100%={s_2/n_sr_2*100}%')
print(f's3/n_sr_3*100%={s_3/n_sr_3*100}%')

for i in range(len(C1)):
    if (-s_3 + n_sr_3 <= C1[i] <= s_3+ n_sr_3):
        C4.append(1)
    if -2*s_3 + n_sr_3 <= C1[i] <= 2*s_3 + n_sr_3:
        D.append(1)
print(f'Доля случаев+-s_3:{sum(C4)/100*100}%')
print(f'Доля случаев+-2*s_3:{sum(D)/100*100}%')

s_n1 = s_1/np.sqrt(400)
e_n1 = s_n1/n_sr_1*100
e_n1_t = 1/np.sqrt(n_sr_1*400)*100

s_n2 = s_2/np.sqrt(200)
e_n2 = s_n2/n_sr_2*100
e_n2_t = 1/np.sqrt(n_sr_2*200)*100

s_n3 = s_3/np.sqrt(100)
e_n3 = s_n3/n_sr_3*100
e_n3_t = 1/np.sqrt(n_sr_3*100)*100

print(f's_n1 = {s_n1}, e_n1 = {e_n1}, e_n1_t = {e_n1_t}')
print(f's_n2 = {s_n2}, e_n2 = {e_n2}, e_n2_t = {e_n2_t}')
print(f's_n3 = {s_n3}, e_n3 = {e_n3}, e_n3_t = {e_n3_t}')
###################__CALCULATIONS__####################################################

##################__UNNECESERY_PART__##################
p = A3[(B3.index(max(B3)))]
x = np.linspace(0, 30, 1000)
y = 1/(s_1*np.sqrt(2*math.pi))*math.e**(-(x-p)**2/(2*s_1**2))
ax3.plot(x,y,color='red',linestyle='--')
#ax3.plot(x1,y1,color='orange',linestyle='--')

p = B[C.index(max(C))]
x = np.linspace(5, 45, 2000)
y = 1/(s_2*np.sqrt(2*math.pi))*math.e**(-(x-25)**2/(2*s_2**2))
ax1.plot(x,y,color='red',linestyle='--')

C3 = []
C4 = []
k = 0
for l in range(min(C1), max(C1) + 1):
    for j in range(0, len(C1)):
        if l == C1[j]:
            k += 1
    C3.append(k)
    k = 0
for i in range(min(C1),max(C1)+1):
    C4.append(i)

s = sum(C3)
for i in range(0, len(C3)):
    C3[i] = C3[i]/s
p = C4[C3.index(max(C3))]

x = np.linspace(30, 80, 2000)
y = 1/(s_3*np.sqrt(2*math.pi))*math.e**(-(x-p)**2/(2*s_3**2))
ax2.plot(x,y,color='red',linestyle='--')
######################################################
from scipy.special import factorial
x = np.arange(0,30)
y = np.exp(-n_sr_1)*(n_sr_1**x)/factorial(x)
ax3.plot(x,y,color='orange',linestyle='--')

y = np.exp(-n_sr_2)*(n_sr_2**x)/factorial(x)
ax1.plot(x,y,color='orange',linestyle='--')

x = np.arange(20,80)
y = np.exp(-n_sr_3)*(n_sr_3**x)/factorial(x)
ax2.plot(x,y,color='orange',linestyle='--')

##################__UNNECESERY_PART__##################
plt.savefig('114_graphic', dpi = 300)
plt.show()